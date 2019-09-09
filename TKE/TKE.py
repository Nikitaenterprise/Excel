from copy import copy
import sys
sys.path.append("..")

import openpyxl
import win32com.client

from Modules.ExcelBook import ExcelBook
from Modules.Header import Header


class TKE:

    def __init__(self, wbName: str):
        self.wb = ExcelBook(wbName, data_only=True)
        self.header = Header(wbName, "A1:BU9")
        self.numberOfRows = self.wb.ws.max_row

    def save(self, name: str):
        self.wb.save(name)
        return

    def start(self):

        #Set 1 and 'договир э' to those companies who have a restructurization contract
        rangeIter = "O12" + ":" + "O" + str(self.numberOfRows)
        for cells in self.wb.ws[rangeIter]:
            for cell in cells:
                if cell.value != "" and cell.value != None:
                    self.wb.ws[str("AM")+str(cell.row)] = str("договір є")
                    if self.wb.ws[str("AS")+str(cell.row)] != "":            # This check needs for empty cell not to be filled
                        self.wb.ws[str("AS")+str(cell.row)] = 1

        #Opens ysterday workbook and copies one specific column to current 
        #workbook shifting other columns
        #fillColor = openpyxl.styles.PatternFill(start_color="cdffcd",
        #                                        fill_type="solid"
        #                                        ) #Set fill color green
        #fillBlue = openpyxl.styles.PatternFill(start_color="9acdff",
        #                                        fill_type="solid"
        #                                        ) #Set fill color blue

        #font = openpyxl.styles.Font(name="Arial",sz=9)      # Set font
        #border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin",
        #                                color='000000'),
        #                                right=openpyxl.styles.Side(border_style="thin",
        #                                color='000000'),
        #                                top=openpyxl.styles.Side(border_style="thin",
        #                                color='000000'),
        #                                bottom=openpyxl.styles.Side(border_style="thin",
        #                                color='000000')
        #                                )                       # Set border
        self.save("ggg.xlsx")
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open("C:\Code\Python\Excel\ggg.xlsx")
        sheet = workbook.Worksheets(1)
        rangeObj = sheet.Range("AS1:AS2")
        rangeObj.EntireColumn.Insert()
        del rangeObj

        wbFromYesterday = ExcelBook("./second.xlsx", data_only=True)

        excelYesterday = win32com.client.Dispatch("Excel.Application")
        excelYesterday.Visible = False
        workbookYesterday = excelYesterday.Workbooks.Open("C:\Code\Python\Excel\second.xlsx") # Right path
        sheetYesterday = workbookYesterday.Worksheets(1)
        sheetYesterday.Range("AN1:AN"+str(wbFromYesterday.ws.max_row)).Unmerge()
        sheet.Range("AS1:AS"+str(self.numberOfRows)).Unmerge()
        rangeObjYesterday = sheetYesterday.Range("AN1:AN"+str(wbFromYesterday.ws.max_row)).Copy()
        sheet.Paste(sheet.Range("AS1:AS"+str(self.numberOfRows)))
        workbookYesterday.Close()
        workbook.Save()
        workbook.Close()
        excelYesterday.Quit()
        excel.Quit()


        columnNumber = openpyxl.utils.column_index_from_string("AS")
        if self.numberOfRows == wbFromYesterday.ws.max_row:                    #TODO make the right check
            rangeIter = "AS1" + ":" + "BU" + str(self.numberOfRows)
            self.wb.ws.move_range(rangeIter, rows=0, cols=1)                
            # for cell in wbFromYesterday.ws["AN"]:
            #     if cell.value == "" or cell.value == None:
            #         self.wb.ws.cell(row = cell.row, 
            #                     column = columnNumber,
            #                     ).fill = fillBlue
            #     else:
            #         self.wb.ws.cell(row = cell.row, 
            #                         column = columnNumber, 
            #                         value = cell.value
            #                         )
            #         self.wb.ws.cell(row = cell.row, 
            #                         column = columnNumber,
            #                         ).font = font
            #         self.wb.ws.cell(row = cell.row, 
            #                         column = columnNumber,
            #                         ).border = border
            #         self.wb.ws.cell(row = cell.row, 
            #                         column = columnNumber,
            #                         ).fill = fillColor
        else:
            raise Exception("Different number of rows in both docs\n. The first has: {}".format(self.wb.ws.max_row))
        
        # Copy names of hidden columns
        self.listOfHiddenColumns = []
        for column in range(1, wbFromYesterday.ws.max_column):
            self.listOfHiddenColumns.append(wbFromYesterday.ws.column_dimensions[openpyxl.utils.get_column_letter(column)].hidden)
        
        listOfRangedCells = wbFromYesterday.ws.merged_cells.ranges
        wbFromYesterday.close()

        #Transfer data from 'поточний лимит' to 'попередний лимит'
        rangeIter1 = "BO10" + ":" + "BU" + str(self.numberOfRows)
        rangeIter2 = "BB10" + ":" + "BH" + str(self.numberOfRows)
        for cells1, cells2 in zip(self.wb.ws[rangeIter1], self.wb.ws[rangeIter2]):
            for cell1, cell2 in zip(cells1, cells2):
                if cell1.row == cell2.row:
                    self.wb.ws.cell(column=cell2.column, row=cell2.row, value=cell1.value)
        
        #Multiply cells in column 'план на декаду' by 3
        rangeIter = "AU10" + ":" + "BA" + str(self.numberOfRows)
        for cells in self.wb.ws[rangeIter]:
            for cell in cells:
                if cell.value != None:
                    self.wb.ws.cell(column=cell.column, row=cell.row, value=cell.value*3)
        
        #Set 'план э' to those rows wich have 0`s in both columns with conditions
        list1 = self.wb.getListOfCellsWithCriteria("AS", 0)             # Check the range
        list2 = self.wb.getListOfCellsWithCriteria("AT", 0)             # Check the range
        
        for cell1 in list1:
            if cell1.value == 0:
                for cell2 in list2:
                    if cell2.value == 0 and cell1.row == cell2.row:
                        self.wb.ws[str("AU")+str(cell1.row)] = str("план є")
                        self.wb.ws[str("AV")+str(cell1.row)] = ""
                        self.wb.ws[str("AW")+str(cell1.row)] = ""
                        self.wb.ws[str("AX")+str(cell1.row)] = ""
                        self.wb.ws[str("AY")+str(cell1.row)] = ""
                        self.wb.ws[str("AZ")+str(cell1.row)] = ""
                        self.wb.ws[str("BA")+str(cell1.row)] = ""

        # Find the difference between columns with 'план на декаду' and 'поточний лимит'
        for row in range(10, self.numberOfRows):
            if self.wb.ws[str("AU")+str(row)].value != "план є" and self.wb.ws[str("AU")+str(row)].value != None:
                dx = self.wb.ws[str("BO")+str(row)].value-self.wb.ws[str("AU")+str(row)].value
                if dx > 1E-4 or dx < -1E-4:                                                         # Check range!!!
                    self.wb.ws.cell(column=openpyxl.utils.column_index_from_string(str("BW")),
                                    row=row,
                                    value=dx
                                    )

        #self.wb.unmerge()
        return

    def hideColumns(self):
        for column in range(1, self.wb.ws.max_column):
            if column < len(self.listOfHiddenColumns):
                if self.listOfHiddenColumns[column] == True:
                    self.wb.ws.column_dimensions[openpyxl.utils.get_column_letter(column)].hidden = True
        return

        
first = TKE("./first.xlsx")
first.start()  
first.hideColumns()
first.save("./out.xlsx")
