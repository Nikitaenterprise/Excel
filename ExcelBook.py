import openpyxl


class ExcelBook:

    def __init__(self, name: str, data_only=False):
        self.wb = openpyxl.load_workbook(name, data_only=data_only)
        self.ws = self.wb[self.wb.sheetnames[0]]

    def save(self, name: str):
        """Closes file and save it to project root dir
        with a given name

        Keyword arguments:
        name -- name of saved file
        """
        self.wb.save(name)
        self.wb.close()
        return
    
    def close(self):
        """Closes file without saving"""
        self.wb.close()
        return 

    def generator(self, range: str):
        """Generator through range in sheet.
        Returns cell
        """
        for cells in self.ws[range]:
            for cell in cells:
                yield cell

    def setCellsInColumnByRowCoord(self, row: int, column: str, value):
        """Finds values in one column by row coordinate
        and then set value of that cell

        Keyword arguments:
        row -- row number
        column -- column in which the searh would happend
        value -- this would be set to the cell
        """
        for cells in self.ws[column]:
            for cell in cells:
                if cell.row == row:
                    cell.value = value
        return

    def setCellsInRowByColumnCoord(self, row: int, column: str, value):
        """Finds values in one row by column coordinate
        and then set value of that cell

        Keyword arguments:
        row -- row number
        column -- column in which the searh would happend
        value -- this would be set to the cell
        """
        for cells in self.ws[str(row)]:
            for cell in cells:
                if cell.column == openpyxl.utils.column_index_from_string(column):
                    cell.value = value
        return

    def findCellByStr(self, value):
        """Finds first cell by searchin in whole sheet 
        the target value

        Keyword argument:
        value -- searching value (str, int, ...)
        """
        for cells in self.ws:
            for cell in cells:
                if cell.value == value:
                    return cell
        return Null

    def getListOfCellsWithCriteria(self, range: str, criteria):
        """Returns list of cells with values equal to criteria

        Keyword argument:
        range -- search range, could be with or 
        without ':' (I22:J22, or I)
        criteria -- search criteria in cells values
        """
        listOfCells = []

        if  range.isdigit() == True:
            for cells in self.ws[range]:
                for cell in cells:
                    if cell.value == criteria:
                        listOfCells.append(cell)
        elif range.isdigit() == False:
            for cell in self.ws[range]:
                if cell.value == criteria:
                    listOfCells.append(cell)
        return listOfCells
