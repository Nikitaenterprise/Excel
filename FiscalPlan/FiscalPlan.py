import os
import datetime
from enum import Enum


import openpyxl
import win32com.client


from src.excel import ExcelBook, hasNumbers

class FiscalPlan:
    
    def __init__(self, dir: str):
        self.day = datetime.datetime.today().day
        self.weekday = datetime.datetime.today().weekday()
        self.checkIsDirectoryIsReady(dir)
    
    def checkIsDirectoryIsReady(self, path: str):
        numberOfFiles = self.scanDirectory(path)
        # Check the dir for needed files
        while True:
            if numberOfFiles == 6:
                break
            if numberOfFiles > 6:
                print("Слишком много экселевских файлов в папке")
                print("Должно быть ровно 6")
                print("Программа пробует удалить ненужные")
                self.deleteFiles(False)
            numberOfFiles = self.scanDirectory(path)

        try:
            self.fiscalPlan
            self.SBUT
            self.PAT
            self.todayCash
            self.lastYearCash
            self.TEZ
        except AttributeError:
            print("Не хватает файлов для работы. Проверьте директорию " + str(path))
            print(self.instructionMessage())
            input()
            exit()
        

    def scanDirectory(self, path: str):
        """Scans the directory with os.walk() for excel files
        and set class excel book variables for folowing work
        """
        print(os.path.abspath(path))
        numberOfFiles = 0
        # r=root, d=directories, f = files
        for r, d, f in os.walk(path):
            for file in f:
                if ".xls" in file or ".xlsx" in file:
                    numberOfFiles += 1
                    print(file)
                    if "Прогнозне надходження" in file:
                        self.fiscalPlan = ExcelBook(path+"\\"+file, read=False)
                    elif "ЗБУТ" in file:
                        self.SBUT = ExcelBook(path+"\\"+file, read=False, keep_vba=True)
                    elif "ПАТ" in file:
                        self.PAT = ExcelBook(path+"\\"+file, read=False, keep_vba=True)
                    elif "НаКР" in file:
                        # Check for date in filename
                        if hasNumbers(file):
                            # Searching for last year money
                            if "-" in file:
                                self.lastYearCash = ExcelBook(path+"\\"+file, read=False)
                            # Searching for yesterday money excel file
                            elif str(self.day - 1) in file:
                                self.todayCash = ExcelBook(path+"\\"+file, read=False)
                            else:
                                print("Будьте осторожны, программа использует файл с деньгами с неправильной датой")
                                self.todayCash = ExcelBook(path+"\\"+file, read=False)
                    elif "ТЕЦ" in file:
                        self.TEZ = ExcelBook(path+"\\"+file, read=False, worksheet=2)
        return numberOfFiles

    def instructionMessage(self):
        msg = """Файлы, нужные для работы: 
            1. _Прогнозне надходження коштів ... 
            2. Всі Категорії. (без спожив. за ...)_ЗБУТ
            3. Всі Категорії. (без спожив. за ...)_ПАТ
            4. ТЕЦ
            5. НадходженняНаКР_ 2 файла: вчерашние деньги и деньги за прошлый год
                Файл со вчерашними деньгами обязан содержать в названии вчерашнюю
                дату, например, если сегодня 13 сентября, пятница, то файл должен называться
                НадходженняНаКР_12 или 12_НадходженняНаКР_ 
                (в этом файле, за прошлый день, НЕЛЬЗЯ использовать символ тире "-"), и т.д.
                Второй файл содержит платежи за период с начала месяца по 12 сентября (как в предыдущем примере), 
                за прошлый год. Имя такого файло должно ОБЯЗАТЕЛЬНО содержать тире "-", например
                НадходженняНаКР_1-12 или НадходженняНаКР(1-12)_
            Итого: 6 экселевских файлов
            После исправления запустите программу заново. Сейчас программа завершит работу
            Нажмите любую клавишу а затем Enter
            """
        return msg

    def run(self):
        self.todayCash.readExcelFile()
        print("1: " + str(self.populationAndReligion(self.todayCash)/1000000))
        print("2: " + str(self.teploenergy(self.todayCash)/1000000))
        print("3: " + str(self.directContractIndustryEE(self.todayCash)/1000000))
        print("4: " + str(self.directContractIndustryPR(self.todayCash)/1000000))

        self.lastYearCash.readExcelFile()
        print("1: " + str(self.populationAndReligion(self.lastYearCash)/1000000))
        print("2: " + str(self.teploenergy(self.lastYearCash)/1000000))
        print("3: " + str(self.directContractIndustryEE(self.lastYearCash)/1000000))
        print("4: " + str(self.directContractIndustryPR(self.lastYearCash)/1000000))

        self.deleteFiles()
        return

    def deleteFiles(self, programmIsDone=True):
        """Deletes all created files with .xlsx extension
        """
        # If programm has daone its work then close files
        if programmIsDone == True:
            try:
                self.closeFiles()
            except:
                print("Программа не смогла закрыть экселевские файлы")

        fileNameWithPathWithoutExtensionTEZ = os.path.splitext(self.TEZ.fileNameWithPath)[0]
        fileNameWithPathWithoutExtensionPAT = os.path.splitext(self.PAT.fileNameWithPath)[0]
        fileNameWithPathWithoutExtensionSBUT = os.path.splitext(self.SBUT.fileNameWithPath)[0]
        fileNameWithPathWithoutExtensionTodayCash = os.path.splitext(
                                                        self.todayCash.fileNameWithPath)[0]
        fileNameWithPathWithoutExtensionLastYearCash = os.path.splitext(
                                                        self.lastYearCash.fileNameWithPath)[0]
        
        numberOfDeletedFiles = 0
        try:
            os.remove(fileNameWithPathWithoutExtensionTEZ + ".xlsx")
        except FileNotFoundError:
            numberOfDeletedFiles += 1
        try:
            os.remove(fileNameWithPathWithoutExtensionPAT + ".xlsx")
        except FileNotFoundError: 
           numberOfDeletedFiles += 1
        try:
            os.remove(fileNameWithPathWithoutExtensionSBUT + ".xlsx")
        except FileNotFoundError:
            numberOfDeletedFiles += 1
        try:
            os.remove(fileNameWithPathWithoutExtensionTodayCash + ".xlsx")
        except FileNotFoundError:
            numberOfDeletedFiles += 1
        try:
            os.remove(fileNameWithPathWithoutExtensionLastYearCash + ".xlsx")
        except FileNotFoundError:
            numberOfDeletedFiles += 1
        return

    def end(self):
        self.closeFiles()
        return

    def closeFiles(self):
        """Closes files without saving
        """
        self.PAT.close()
        self.SBUT.close()
        self.TEZ.close()
        self.todayCash.close()
        self.lastYearCash.close()
        return

    class ColumnsNames(Enum):
        COMPANY = 0
        CATEGORY = 0
        CONTRACT = 1
        CASH = 2

    def initColumnIndecesForCashBook(self):
        
        return
    
    def populationAndReligion(self, cashWB: ExcelBook):
        """Finds sum of cash from population and religion
        """
        try:
            # Column C contain names and categories
            populationRow = cashWB.getFirstCellByCriteria("1.2. Населення", "C").row
            # Column J contain cash
            populationColumn = openpyxl.utils.column_index_from_string(str("J"))
            populationCash = cashWB.ws.cell(column=populationColumn, row=populationRow).value
        except:
            print("Нет категории: Населення")
            populationCash = 0

        try:
            # Column C contain names and categories
            religionRow = cashWB.getFirstCellByCriteria("Релігійні організації", "C").row
            # Column J contain cash
            religionColumn = populationColumn # the same column
            religionCash = cashWB.ws.cell(column=religionColumn, row=religionRow).value
        except:
            print("Нет категории: Релігійні організації")
            religionCash = 0

        return populationCash + religionCash

    def teploenergy(self, cashWB: ExcelBook):
        """Finds sum of cash from PSO users (ТЕ, КП, БО, РО)
        """
        try:
            # Column C contain names and categories
            teploenergyRow = cashWB.getFirstCellByCriteria("3.2. Теплоенергетика за прямими договорами", "C").row
            # Column J contain cash
            teploenergyColumn = openpyxl.utils.column_index_from_string(str("J"))
            teploenergyCash = cashWB.ws.cell(column=teploenergyColumn, row=teploenergyRow).value
        except:
            print("Нет категории: Теплоенергетика за прямими договорами")
            teploenergyCash = 0

        try:
            # Column C contain names and categories
            kyivEnergoNotEeRow = cashWB.getFirstCellByCriteria("Енергетичні підприємства м.Києва", "C").row
            # Column J contain cash
            kyivEnergoNotEeColumn = teploenergyColumn # the same column
            kyivEnergoNotEeCash = 0
            
            # Column E contain names of contracts wich were concluded with companies 
            kyivEnergoNotEeColumnWithNameOfContract = openpyxl.utils.column_index_from_string(str("E"))
            # Here, the contracts wich not contain EE are calculated
            for i in range(2, 6):
                kyivEnergoNotEeRow += i
                contractName = cashWB.ws.cell(
                                column=kyivEnergoNotEeColumnWithNameOfContract, 
                                row=kyivEnergoNotEeRow).value
                cashValue = cashWB.ws.cell(
                                column=kyivEnergoNotEeColumn, 
                                row=kyivEnergoNotEeRow).value
                if "ЕЕ" in contractName:
                    self.kyivEnergoEeContractCash = cashValue
                elif "ЕЕ" not in contractName:
                    kyivEnergoNotEeCash += cashValue
        except:
            print("Нет категории: Енергетичні підприємства м.Києва")
            kyivEnergoNotEeCash = 0

        return teploenergyCash + kyivEnergoNotEeCash

    def directContractIndustryEE(self, cashWB: ExcelBook):
        """Finds sum of cash from direct contract with
        industries (EE) and TEZ companies
        """
        try:
            self.TEZ.readExcelFile()
            self.TEZ.unmerge()
            # Column C contain names and categories
            industryEeRow = cashWB.getFirstCellByCriteria("2.2. Промисловість за прямими договорами", "C").row
            # Column J contain cash
            industryEeColumnWithCash = openpyxl.utils.column_index_from_string(str("J"))
            # Column E contain names of contracts wich were concluded with companies
            industryEeColumnWithNameOfContracts = openpyxl.utils.column_index_from_string(str("E"))
            # Column C contain names and categories
            industryEeColumnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(str("C"))

            industryEeCash = 0
            TEZCash = 0
            while True:
                industryEeRow += 1
                categoryOrCompanyName = cashWB.ws.cell(
                                column=industryEeColumnWithNameOfCompanyOrCategory, 
                                row=industryEeRow).value
                if categoryOrCompanyName == "Всього по теплоенергетиці":
                    break

                contractName = cashWB.ws.cell(
                                column=industryEeColumnWithNameOfContracts, 
                                row=industryEeRow).value
                if "ЕЕ" in contractName:
                    industryEeCash += cashWB.ws.cell(
                                column=industryEeColumnWithCash, 
                                row=industryEeRow).value                    
                else:
                    # Search the ТЕЦ.xlsx file for companies overlaping
                    # and if so, cash of this companies would be added
                    try:
                        # Get cell from ТЕЦ.xlsx with company name that is equal to 
                        # company name in НадходженняНаКР_.xlsx
                        cell = self.TEZ.getFirstCellByCriteria(str(categoryOrCompanyName), "D")
                        if cell == None:
                            raise AttributeError(cell)

                        cellValue = cell.value
                    except AttributeError:
                        # If cell have None type then there is no such company in 
                        # ТЕЦ.xlsx and thats why cellValue should be equal to 0
                        cellValue = 0

                    if cellValue != 0:
                        TEZCash += cashWB.ws.cell(
                                column=industryEeColumnWithCash, 
                                row=industryEeRow).value
        except:
            print("Нет категории: Промисловість за прямими договорами (ЕЕ)")
            industryEeCash = 0
            TEZCash = 0

        try:
            # Check is this vatiable is created
            # in teploenergy()
            kyivEnergoEeContractCash
        except:
            print("Нет категории: Енергетичні підприємства м.Києва (ЕЕ)")
            kyivEnergoEeContractCash = 0

        return industryEeCash + TEZCash + kyivEnergoEeContractCash

    def directContractIndustryPR(self, cashWB: ExcelBook):
        """Finds cash from direct contract with
        industries (PR) without cash from TEZ companies and
        SBUT companies and PAT companies
        """
        try:
            self.PAT.readExcelFile()
            self.PAT.unmerge()
            self.SBUT.readExcelFile()
            self.SBUT.unmerge()

            # Column C contain names and categories
            industryPrRow = cashWB.getFirstCellByCriteria("2.2. Промисловість за прямими договорами", "C").row
            # Column J contain cash
            industryPrColumn = openpyxl.utils.column_index_from_string(str("J"))
            # Column E contain names of contracts wich were concluded with companies
            industryPrColumnWithNameOfContracts = openpyxl.utils.column_index_from_string(str("E"))
            # Column C contain names and categories
            industryPrColumnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(str("C"))

            industryPrCash = 0
            naftogazTradingCash = 0
            while True:
                industryPrRow += 1
                categoryOrCompanyName = cashWB.ws.cell(
                                column=industryPrColumnWithNameOfCompanyOrCategory, 
                                row=industryPrRow).value

                if categoryOrCompanyName == "Всього по теплоенергетиці":
                    break

                contractName = cashWB.ws.cell(
                                column=industryPrColumnWithNameOfContracts, 
                                row=industryPrRow).value
                if "ПР" in contractName:
                    # Search in:
                    # ТЕЦ.xlsx, 
                    # Всі Категорії. (без спожив. за )_ЗБУТ.xlsx, 
                    # Всі Категорії. (без спожив. за )_ПАТ.xlsx 
                    # files for companies overlaping and if so, 
                    # cash of this companies wouldn`t be calculated
                    try:
                        cellTEZ = self.TEZ.getFirstCellByCriteria(str(categoryOrCompanyName), "D")
                        cellPAT = self.PAT.getFirstCellByCriteria(str(categoryOrCompanyName), "C")
                        cellSBUT = self.SBUT.getFirstCellByCriteria(str(categoryOrCompanyName), "C")
                        if cellTEZ != None or cellPAT != None or cellSBUT != None:
                            continue
                        # if cellTEZ == None or cellPAT == None or cellSBUT == None:
                        #     raise AttributeError()
                    except AttributeError:
                        continue
                        
                    industryPrCash += cashWB.ws.cell(
                            column=industryPrColumn, 
                            row=industryPrRow).value  

                if "НАФТОГАЗ ТРЕЙДИНГ" in categoryOrCompanyName:
                    naftogazTradingCash += cashWB.ws.cell(
                            column=industryPrColumn, 
                            row=industryPrRow).value 
        except:
            print("Нет категории: Промисловість за прямими договорами (ПР)")
            industryPrCash = 0
            naftogazTradingCash = 0

        try:
            # Column J contain cash
            energoGenerationColumn = openpyxl.utils.column_index_from_string(str("J"))
            # Column C contain company name or category
            energoGenerationColumnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(str("C"))
            # Column E contain names of contracts wich were concluded with companies
            energoGenerationColumnWithNameOfContracts = openpyxl.utils.column_index_from_string(str("E"))
            # Column C contain company name or category
            energoGenerationRow = cashWB.getFirstCellByCriteria("Енергогенеруючі компанії", "C").row
            
            while True:
                energoGenerationRow += 1
                categoryOrCompanyName = cashWB.ws.cell(
                                column=energoGenerationColumnWithNameOfCompanyOrCategory, 
                                row=energoGenerationRow).value
                if categoryOrCompanyName == "Релігійні організації":
                    break
                
                contractName = cashWB.ws.cell(
                                column=energoGenerationColumnWithNameOfContracts, 
                                row=energoGenerationRow).value
                if "ПР" in contractName:
                    energoGenerationCash += cashWB.ws.cell(
                                column=energoGenerationColumn, row=energoGenerationRow)
        except:
            print("Нет категории: Енергогенеруючі компанії (ПР)")
            energoGenerationCash = 0


        print('Деньги от ТОВ "ГАЗОПОСТАЧАЛЬНА КОМПАНІЯ "НАФТОГАЗ ТРЕЙДИНГ" ' + str(naftogazTradingCash/1000000))
        return industryPrCash + energoGenerationCash

    def addToSummaryFile(self):
        self.fiscalPlan.readExcelFile()
        rowWithDates = 5
        columnWithDates = openpyxl.utils.column_index_from_string(str("B"))
        listOfHeaders = getHeadersOfAllColumns()
        while True:
            cellValue = self.fiscalPlan.ws.cell(column=columnWithDates, row=rowWithDates).value
            
            isThereAFactCell = False
            if "факт" in cellValue:
                isThereAFactCell = True
            isThereAPlanCell = False
            if "план" in cellValue:
                isThereAPlanCell = True
            if isThereAFactCell == True and isThereAPlanCell == True:
                columnWithDates += 1
                continue
            #elif isThereAFactCell == True and isThereAPlanCell == False:
    def ppp(self):

        self.fiscalPlan.readExcelFile()
        rowWithDates = 5
        columnWithDates = openpyxl.utils.column_index_from_string(str("B"))
        self.fiscalPlan.initHeader("B4:AH5")
        listOfHeaders = self.fiscalPlan.getHeadersOfAllColumns()



        excelApp, wb = self.fiscalPlan.readFileWithPyWin()
        
        #ws = wb.Worksheets("вересень")
        ws = wb.ActiveSheet
        coord = openpyxl.utils.coordinate_to_tuple("H6")
        print(ws.Cells(coord[0], coord[1]).Value)
        range = ws.Range("I1:I10")
        range.EntireColumn.Insert()

        excelApp.Quit()


if __name__ == "__main__":
    print("I`m FiscalPlan.py file")
    #run tests
