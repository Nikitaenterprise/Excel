import os
import datetime

import openpyxl
import win32com.client

from src.manager import *


class FiscalPlan:

    def __init__(self, dir: str):
        self.mng = Manager(os.path.abspath(dir))
        self.numberOfFilesToStart = 6
        self.checkIfDirectoryIsReady(dir)

    def checkIfDirectoryIsReady(self, path: str):
        self.mng.addFilesInDir()
        
        self.fiscalPlan = self.mng.getFile("Прогнозне надходження", extension=".xlsx")
        self.fiscalPlan.shouldBeDeleted = False
        self.mng.getFile("ЗБУТ")
        self.mng.getFile("ПАТ")
        self.mng.getFile("ТЕЦ")
        self.mng.getFile("НаКР")
        self.mng.getFile("НаКР")

        self.mng.deleteUnCalledFiles()               
        self.mng.allFromXlsToXlsx()

        try:
            self.SBUT = self.mng.getFile("ЗБУТ", extension=".xlsx")
            self.PAT = self.mng.getFile("ПАТ", extension=".xlsx")
            self.TEZ = self.mng.getFile("ТЕЦ", extension=".xlsx")
            
            cash = []
            cash.append(self.mng.getFile("НаКР", extension=".xlsx"))
            cash.append(self.mng.getFile("НаКР", extension=".xlsx"))
            for cashFile in cash:

                try:
                    fileName = cashFile.fileName
                except AttributeError:
                    print("Проблема с файлом НаКР, возможно, он отсутствует")
                    raise WindowsError

                if hasNumbers(fileName):
                    day = datetime.datetime.today().day
                    # Searching for last year money
                    if "-" in fileName:
                        self.lastYearCash = cashFile
                    # Searching for yesterday money excel file
                    elif str(day - 1) in fileName:
                        self.todayCash = cashFile
                    else:
                        print(
                            "Будьте осторожны, программа использует файл с деньгами с неправильной датой")
                        self.todayCash = cashFile

                if self.mng.getNumberOfFiles() != self.numberOfFilesToStart:
                    raise AttributeError

        except AttributeError:
            print("Не хватает файлов для работы. Проверьте директорию " + str(path))
            msg = """Файлы, нужные для работы: 
            1. _Прогнозне надходження коштів ... 
            2. Всі Категорії. (без спожив. за ...)_ЗБУТ
            3. Всі Категорії. (без спожив. за ...)_ПАТ
            4. ТЕЦ
            5. НадходженняНаКР_ 2 файла: вчерашние деньги и деньги за прошлый год
                Файл со вчерашними деньгами обязан содержать в названии вчерашнюю
                дату, например, если сегодня 13 сентября, пятница, то файл должен называться
                НадходженняНаКР_12 или 12_НадходженняНаКР_ 
                (в этом файле, за прошлый день, НЕЛЬЗЯ использовать символ тире "-"), и т.д.
                Второй файл содержит платежи за период с начала месяца по 12 сентября (как в предыдущем примере), 
                за прошлый год. Имя такого файло должно ОБЯЗАТЕЛЬНО содержать тире "-", например
                НадходженняНаКР_1-12 или НадходженняНаКР(1-12)_
            Итого: 6 экселевских файлов
            После исправления запустите программу заново. Сейчас программа завершит работу
            Нажмите любую клавишу а затем Enter
            """
            print(msg)
            input()
            exit()

    def deleteFiles(self, programmIsDone=True):
        """Deletes all created files with .xlsx extension
        """
        # If programm has daone its work then close files
        if programmIsDone == True:
            try:
                self.PAT.close()
                self.SBUT.close()
                self.TEZ.close()
                self.todayCash.close()
                self.lastYearCash.close()
            except:
                print("Программа не смогла закрыть экселевские файлы")
            try:
                self.mng.deleteClosedFiles()
            except FileNotFoundError:
                print("Программа не смогла удалить файлы после работы")
        return

    def run(self):
        print("\n\n\n")
        print("Обработка файла с деньгами за предыдущий")
        print("банковский день")
        self.todayCash.open()
        today = []
        today.append(self.populationAndReligion(self.todayCash)/1000000)
        today.append(self.teploenergy(self.todayCash)/1000000)
        today.append(self.directContractIndustryEE(self.todayCash)/1000000)
        listWithPR = self.directContractIndustryPR(self.todayCash)
        listWithVTV = self.additionalIncome(self.todayCash)
        # PR money
        today.append(listWithPR[0]/1000000)
        # Naftogaz money
        today.append(listWithPR[1]/1000000)
        # UKRTRANSGAS VTV
        today.append(listWithVTV[1]/1000000)
        # Additional income 
        today.append(listWithVTV[0]/1000000)
        
        # Add TEZ money from PR category to EE category
        today[2] += self.tezCompaniesMoneyFromPr/1000000

        # List insides:
        # [
        #       Населення; 
        #       Теплокомуненерго по договорах ТЕ,БО, КП, РО; 
        #       Теплокомуненерго по договорах ЕЕ;
        #       Промислові підприємства; 
        #       Нафтогаз Трейдинг;
        #       УКРТРАНСГАЗ (ВТВ);
        #       Додаткові надходження
        # ]

        print("Деньги за сегодня")
        print("\tНаселення", today[0])
        print("\tТеплокомуненерго по договорах ТЕ,БО, КП, РО", today[1])
        print("\tТеплокомуненерго по договорах ЕЕ", today[2])
        print("\tПромислові підприємства", today[3])
        print("\tНафтогаз Трейдинг", today[4])
        print("\tУКРТРАНСГАЗ", today[5])
        print("\tДодаткові надходження", today[6])
        
        print("\n\n\n")
        print("Обработка файла с деньгами за месяц")
        print("в прошлом году")
        self.lastYearCash.open()
        lastYear = []
        lastYear.append(self.populationAndReligion(self.lastYearCash)/1000000)
        lastYear.append(self.teploenergy(self.lastYearCash)/1000000)
        lastYear.append(self.directContractIndustryEE(self.lastYearCash)/1000000)
        listWithPR = self.directContractIndustryPR(self.lastYearCash)
        listWithVTV = self.additionalIncome(self.lastYearCash)
        # PR money
        lastYear.append(listWithPR[0]/1000000)
        # Naftogaz money
        lastYear.append(listWithPR[1]/1000000)
        # UKRTRANSGAS VTV
        lastYear.append(listWithVTV[1]/1000000)
        # Additional income 
        lastYear.append(listWithVTV[0]/1000000)

        # Add TEZ money from PR category to EE category
        lastYear[2] += self.tezCompaniesMoneyFromPr/1000000

        print("Деньги за прошлый год")
        print("\tНаселення", lastYear[0])
        print("\tТеплокомуненерго по договорах ТЕ,БО, КП, РО", lastYear[1])
        print("\tТеплокомуненерго по договорах ЕЕ", lastYear[2])
        print("\tПромислові підприємства", lastYear[3])
        print("\tНафтогаз Трейдинг", lastYear[4])
        print("\tУКРТРАНСГАЗ", lastYear[5])
        print("\tДодаткові надходження", lastYear[6])

        self.fillPlan(today, lastYear)
        return

    def populationAndReligion(self, cashWB: File):
        """Finds sum of cash from population and religion
        """
        try:
            # Column C contain names and categories
            populationRow = cashWB.getFirstCellByCriteria(
                "1.2. Населення", "C").row
            # Column J contain cash
            populationColumn = openpyxl.utils.column_index_from_string(
                str("J"))
            populationCash = cashWB.getWs(0).cell(
                column=populationColumn, row=populationRow).value
        except AttributeError:
            print("Нет категории: Населення")
            populationCash = 0

        try:
            # Column C contain names and categories
            religionRow = cashWB.getFirstCellByCriteria(
                "Релігійні організації", "C").row
            # Column J contain cash
            religionColumn = populationColumn  # the same column
            religionCash = cashWB.getWs(0).cell(
                column=religionColumn, row=religionRow).value
        except AttributeError:
            print("Нет категории: Релігійні організації")
            religionCash = 0

        return populationCash + religionCash

    def teploenergy(self, cashWB: File):
        """Finds sum of cash from PSO users (ТЕ, КП, БО, РО)
        """
        try:
            # Column C contain names and categories
            teploenergyRow = cashWB.getFirstCellByCriteria(
                "3.2. Теплоенергетика за прямими договорами", "C").row
            # Column J contain cash
            teploenergyColumn = openpyxl.utils.column_index_from_string(
                str("J"))
            teploenergyCash = cashWB.getWs(0).cell(
                column=teploenergyColumn, row=teploenergyRow).value
        except AttributeError:
            print("Нет категории: Теплоенергетика за прямими договорами")
            teploenergyCash = 0

        try:
            # Column C contain names and categories
            kyivEnergoNotEeRow = cashWB.getFirstCellByCriteria(
                "Енергетичні підприємства м.Києва", "C").row
            # Make one step down in row
            kyivEnergoNotEeRow += 1

            # Column J contain cash
            kyivEnergoNotEeColumn = teploenergyColumn  # the same column
            # Column E contain names of contracts wich were concluded with companies
            kyivEnergoNotEeColumnWithNameOfContract = openpyxl.utils.column_index_from_string(
                str("E"))
            # Column B contain category number
            categoryNumberColumn = openpyxl.utils.column_index_from_string(str("B"))
            
            # Here, the contracts wich not contain EE are calculated
            kyivEnergoNotEeCash = 0
            while True:
                kyivEnergoNotEeRow += 1
                categoryNumber = cashWB.getWs(0).cell(
                        column=categoryNumberColumn,
                        row=kyivEnergoNotEeRow).value
                # Break if category ends
                if categoryNumber != None:
                    break

                contractName = cashWB.getWs(0).cell(
                    column=kyivEnergoNotEeColumnWithNameOfContract,
                    row=kyivEnergoNotEeRow).value
                cashValue = cashWB.getWs(0).cell(
                    column=kyivEnergoNotEeColumn,
                    row=kyivEnergoNotEeRow).value
                if "ЕЕ" in contractName:
                    self.kyivEnergoEeContractCash = cashValue
                elif "ЕЕ" not in contractName:
                    kyivEnergoNotEeCash += cashValue
        except (AttributeError, TypeError):
            print("Нет категории: Енергетичні підприємства м.Києва")
            kyivEnergoNotEeCash = 0

        return teploenergyCash + kyivEnergoNotEeCash

    def directContractIndustryEE(self, cashWB: File):
        """Finds sum of cash from direct contract with
        industries (EE) and TEZ companies
        """
        try:
            self.TEZ.open()
            self.TEZ.unmerge()
            # Column C contain names and categories
            industryEeRow = cashWB.getFirstCellByCriteria(
                "2.2. Промисловість за прямими договорами", "C").row
            # Column J contain cash
            industryEeColumnWithCash = openpyxl.utils.column_index_from_string(
                str("J"))
            # Column E contain names of contracts wich were concluded with companies
            industryEeColumnWithNameOfContracts = openpyxl.utils.column_index_from_string(
                str("E"))
            # Column C contain names and categories
            industryEeColumnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(
                str("C"))
            # Column B contain category number
            categoryNumberColumn = openpyxl.utils.column_index_from_string(str("B"))

            industryEeCash = 0
            TEZCash = 0
            while True:
                industryEeRow += 1
                categoryNumber = cashWB.getWs(0).cell(
                        column=categoryNumberColumn,
                        row=industryEeRow).value
                # Break if category ends
                if categoryNumber != None:
                    break
                
                categoryOrCompanyName = cashWB.getWs(0).cell(
                    column=industryEeColumnWithNameOfCompanyOrCategory,
                    row=industryEeRow).value
                contractName = cashWB.getWs(0).cell(
                    column=industryEeColumnWithNameOfContracts,
                    row=industryEeRow).value
                if "ЕЕ" in contractName:
                    industryEeCash += cashWB.getWs(0).cell(
                        column=industryEeColumnWithCash,
                        row=industryEeRow).value
                else:
                    # Search the ТЕЦ.xlsx file for companies overlaping
                    # and if so, cash of this companies would be added
                    try:
                        # Get cell from ТЕЦ.xlsx with company name that is equal to
                        # company name in НадходженняНаКР_.xlsx
                        cell = self.TEZ.getFirstCellByCriteria(
                            str(categoryOrCompanyName), "D", wsName="Print_Form_1")
                        if cell == None:
                            raise AttributeError(cell)

                        cellValue = cell.value
                    except AttributeError:
                        # If cell have None type then there is no such company in
                        # ТЕЦ.xlsx and thats why cellValue should be equal to 0
                        cellValue = 0

                    if cellValue != 0:
                        TEZCash += cashWB.getWs(0).cell(
                            column=industryEeColumnWithCash,
                            row=industryEeRow).value
        except (AttributeError, TypeError):
            print("Нет категории: Промисловість за прямими договорами (ЕЕ)")
            industryEeCash = 0
            TEZCash = 0

        try:
            # Check is this vatiable is created
            # in teploenergy()
            self.kyivEnergoEeContractCash
        except (AttributeError, UnboundLocalError):
            print("Нет категории: Енергетичні підприємства м.Києва (ЕЕ)")
            self.kyivEnergoEeContractCash = 0

        return industryEeCash + TEZCash + self.kyivEnergoEeContractCash

    def directContractIndustryPR(self, cashWB: File):
        """Finds cash from direct contract with
        industries (PR) without cash from TEZ companies and
        SBUT companies and PAT companies. Also finds money from 
        Naftogaz trading.
        Returns list
        """
        
        try:
            self.PAT.open()
            self.PAT.unmerge()
            self.SBUT.open()
            self.SBUT.unmerge()

            # Column C contain names and categories
            industryPrRow = cashWB.getFirstCellByCriteria(
                "2.2. Промисловість за прямими договорами", "C").row
            # Column J contain cash
            industryPrColumn = openpyxl.utils.column_index_from_string(
                str("J"))
            # Column E contain names of contracts wich were concluded with companies
            industryPrColumnWithNameOfContracts = openpyxl.utils.column_index_from_string(
                str("E"))
            # Column C contain names and categories
            industryPrColumnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(
                str("C"))
            # Column B contain category number
            categoryNumberColumn = openpyxl.utils.column_index_from_string(str("B"))

            industryPrCash = 0
            naftogazTradingCash = 0
            self.tezCompaniesMoneyFromPr = 0
            while True:
                industryPrRow += 1
                categoryNumber = cashWB.getWs(0).cell(
                        column=categoryNumberColumn,
                        row=industryPrRow).value
                # Break if category ends
                if categoryNumber != None:
                    break
                
                categoryOrCompanyName = cashWB.getWs(0).cell(
                    column=industryPrColumnWithNameOfCompanyOrCategory,
                    row=industryPrRow).value
                contractName = cashWB.getWs(0).cell(
                    column=industryPrColumnWithNameOfContracts,
                    row=industryPrRow).value
                if "ПР" in contractName:
                    # Search in:
                    # ТЕЦ.xlsx,
                    # Всі Категорії. (без спожив. за )_ЗБУТ.xlsx,
                    # Всі Категорії. (без спожив. за )_ПАТ.xlsx
                    # files for companies overlaping and if so,
                    # cash of this companies wouldn`t be calculated
                    try:
                        cellTEZ = self.TEZ.getFirstCellByCriteria(
                            str(categoryOrCompanyName), "D", wsName="Print_Form_1")
                        cellPAT = self.PAT.getFirstCellByCriteria(
                            str(categoryOrCompanyName), "C")
                        cellSBUT = self.SBUT.getFirstCellByCriteria(
                            str(categoryOrCompanyName), "C")
                        if cellTEZ != None or cellPAT != None or cellSBUT != None:
                            # If this is TEZ company then add save it`s money
                            # for EE category in directContractIndustryEE()
                            if cellTEZ != None:
                                self.tezCompaniesMoneyFromPr += cashWB.getWs(0).cell(
                                            column=industryPrColumn,
                                            row=industryPrRow).value
                            continue
                    except AttributeError:
                        continue

                    industryPrCash += cashWB.getWs(0).cell(
                        column=industryPrColumn,
                        row=industryPrRow).value

                if "НАФТОГАЗ ТРЕЙДИНГ" in categoryOrCompanyName:
                    naftogazTradingCash += cashWB.getWs(0).cell(
                        column=industryPrColumn,
                        row=industryPrRow).value
        except (AttributeError, TypeError):
            print("Нет категории: Промисловість за прямими договорами (ПР)")
            industryPrCash = 0
            naftogazTradingCash = 0
            self.tezCompaniesMoneyFromPr = 0

        try:
            # Column J contain cash
            energoGenerationColumn = openpyxl.utils.column_index_from_string(
                str("J"))
            # Column C contain company name or category
            energoGenerationColumnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(
                str("C"))
            # Column E contain names of contracts wich were concluded with companies
            energoGenerationColumnWithNameOfContracts = openpyxl.utils.column_index_from_string(
                str("E"))
            
            # Column C contain company name or category
            energoGenerationRow = cashWB.getFirstCellByCriteria(
                "Енергогенеруючі компанії", "C").row
            # Make one step down in row
            energoGenerationRow += 1

            # Column B contain category number
            categoryNumberColumn = openpyxl.utils.column_index_from_string(str("B"))

            energoGenerationCashPR = 0
            energoGenerationCashEE = 0
            while True:
                energoGenerationRow += 1
                categoryNumber = cashWB.getWs(0).cell(
                        column=categoryNumberColumn,
                        row=energoGenerationRow).value
                # Break if category ends
                if categoryNumber != None:
                    break
                
                categoryOrCompanyName = cashWB.getWs(0).cell(
                    column=energoGenerationColumnWithNameOfCompanyOrCategory,
                    row=energoGenerationRow).value
                contractName = cashWB.getWs(0).cell(
                    column=energoGenerationColumnWithNameOfContracts,
                    row=energoGenerationRow).value
                if "ПР" in contractName:
                    energoGenerationCashPR += cashWB.getWs(0).cell(
                            column=energoGenerationColumn, 
                            row=energoGenerationRow).value
                elif "ЕЕ" in contractName:
                    energoGenerationCashEE += cashWB.getWs(0).cell(
                            column=energoGenerationColumn, 
                            row=energoGenerationRow).value
                    
        except (AttributeError, TypeError):
            print("Нет категории: Енергогенеруючі компанії (ПР)")
            energoGenerationCashPR = 0
            energoGenerationCashEE = 0

        self.tezCompaniesMoneyFromPr += energoGenerationCashEE

        return [industryPrCash + energoGenerationCashPR, naftogazTradingCash]

    def additionalIncome(self, cashWB: File):
        """Finds VTV money and UKRTRANSGAZ VTV money
        Returns list
        """
        # Column J contain cash
        cashColumn = openpyxl.utils.column_index_from_string(str("J"))
        try:
            # Column C contain names and categories
            prVatRow = cashWB.getFirstCellByCriteria(
                "2.1. Промисловість через ВАТ", "C").row
            prVatCash = cashWB.getWs(0).cell(
                    column=cashColumn,
                    row=prVatRow).value
        except AttributeError:
            print("Нет категории: Промисловість через ВАТ")
            prVatCash = 0
        try:
            # Column C contain names and categories
            teploVatRow = cashWB.getFirstCellByCriteria(
                "3.1. Теплоенергетика через ВАТ", "C").row
            teploVatCash = cashWB.getWs(0).cell(
                    column=cashColumn,
                    row=teploVatRow).value
        except AttributeError:
            print("Нет категории: Теплоенергетика через ВАТ")
            teploVatCash = 0
        try:
            vtvRow = cashWB.getFirstCellByCriteria(
                "ВТВ та нормовані втрати", "C").row
            vtvCash = cashWB.getWs(0).cell(
                    column=cashColumn,
                    row=vtvRow).value
        except AttributeError:
            print("Нет категории: ВТВ та нормовані втрати")
            vtvCash = 0
        try:
            # Column C contain names and categories
            columnWithNameOfCompanyOrCategory = openpyxl.utils.column_index_from_string(
                str("C"))
            # Column B contain category number
            categoryNumberColumn = openpyxl.utils.column_index_from_string(str("B"))
            transGasVtvCash = 0
            while True:
                vtvRow += 1
                categoryNumber = cashWB.getWs(0).cell(
                        column=categoryNumberColumn,
                        row=vtvRow).value
                # Break if category ends
                if categoryNumber != None:
                    break

                categoryOrCompanyName = cashWB.getWs(0).cell(
                        column=columnWithNameOfCompanyOrCategory,
                        row=vtvRow).value
                if "Оператор ГТС" in categoryOrCompanyName:
                    transGasVtvCash += cashWB.getWs(0).cell(
                            column=cashColumn,
                            row=vtvRow).value
        except UnboundLocalError:
            print("Нет компании: Філія Оператор ГТС України")
            transGasVtvCash = 0
        
        return  [prVatCash + teploVatCash + vtvCash - transGasVtvCash, transGasVtvCash]

    def fillPlan(self, todayMoney: list, lastYearMoney: list):

        self.fiscalPlan.open(data_only=False)

        # Get date frome file name
        newstr = ''.join((ch if ch in '0123456789' else '') \
                    for ch in self.todayCash.fileNameWithoutExtension)
        #listOfNumbers = [int(i) for i in newstr.split()]
        dayInFileName = int(newstr)

        # Find date in header of the excel file 
        cellWithDate = self.fiscalPlan.getFirstCellByCriteria(
                        dayInFileName, "B4:AF4")
        
        # Iterate in excel book in one column (current day) in 5 rows
        for i in range(1, 6):
            self.fiscalPlan.getWs().cell(
                    column=cellWithDate.column,
                    row = cellWithDate.row + i).value = todayMoney[i-1]
        # UKRTRANSGAS(VTV)
        self.fiscalPlan.getWs().cell(
                    column=cellWithDate.column,
                    row = cellWithDate.row + 7).value = todayMoney[5]
        # Additional income
        self.fiscalPlan.getWs().cell(
                    column=cellWithDate.column,
                    row = cellWithDate.row + 8).value = todayMoney[6]

        for i in range(1, 6):
            self.fiscalPlan.getWs().cell(
                    column=openpyxl.utils.column_index_from_string("AH"),
                    row=cellWithDate.row + i).value = lastYearMoney[i-1]
        # UKRTRANSGAS(VTV)
        self.fiscalPlan.getWs().cell(
                    column=openpyxl.utils.column_index_from_string("AH"),
                    row = cellWithDate.row + 7).value = lastYearMoney[5]
        # Additional income
        self.fiscalPlan.getWs().cell(
                    column=openpyxl.utils.column_index_from_string("AH"),
                    row = cellWithDate.row + 8).value = lastYearMoney[6]

        self.fiscalPlan.save(self.fiscalPlan.pathToFile, 
                        self.fiscalPlan.fileNameWithoutExtension)
        self.deleteFiles()


if __name__ == "__main__":
    print("I`m FiscalPlan.py file")
    # run tests
