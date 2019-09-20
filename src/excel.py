import os


import openpyxl
import win32com



def hasNumbers(inputString: str):
    """Checks string for containing numbers
    returns True if string has at least one digit
    """
    return any(char.isdigit() for char in inputString)

def incertColumnWithPyWin(file: File, column: str):
    """Incerts column using pyWin. 
    Function incerts column to the right at first worksheet

    Keyword arguments:
    column -- name of column near what the column would be inserted
    """
    ws = file.getWs(isActiveSheet=True)
    # select column as range object
    if column.isdigit():
        openpyxl.utils.get_column_letter(column)
        newColumn = openpyxl.utils.get_column_letter(column)
        rangeObj = ws.Range(newColumn+str(1)+str(":")+newColumn+str(2))
    else:
        rangeObj = ws.Range(column+str(1)+str(":")+column+str(2))

    rangeObj.EntireColumn.Insert()
    file.save(file.pathToFile, file.fileName)

def setCellsInColumnByRowCoord(file: File, wsNumber=0, row: int, column: str, value):
    """Finds values in one column by row coordinate
    and then set value of that cell

    Keyword arguments:
    row -- row number
    column -- column in which the searh would happend
    value -- this would be set to the cell
    """
    for cells in file.getWs(wsNumber)[column]:
        for cell in cells:
            if cell.row == row:
                cell.value = value
    return

def setCellsInRowByColumnCoord(file: File, wsNumber=0, row: int, column: str, value):
    """Finds values in one row by column coordinate
    and then set value of that cell

    Keyword arguments:
    row -- row number
    column -- column in which the searh would happend
    value -- this would be set to the cell
    """
    for cells in file.getWs(wsNumber)[str(row)]:
        for cell in cells:
            if cell.column == openpyxl.utils.column_index_from_string(column):
                cell.value = value
    return

def getFirstCellByCriteria(file: File, wsNumber=0, criteria, range: str = None):
    """Finds first cell by searchin in whole sheet or in 
    some range the target criteria. Similar to Excel 
    function VLOOKUP (ВПР)

    Keyword arguments:
    value -- searching value (str, int, ...)
    range -- search range (I22:J22, or I), by default set to None
                so it search in whole sheet
    """
    if range == None:
        diapason = file.getWs(wsNumber)
        for cells in diapason:
            for cell in cells:
                if cell.value == criteria:
                    return cell
    elif range != None:
        diapason = file.getWs(wsNumber)[range]
        if hasNumbers(range) == True:
            for cells in diapason:
                for cell in cells:
                    if cell.value == criteria:
                        return cell
        elif hasNumbers(range) == False:
            for cell in diapason:
                if cell.value == criteria:
                    return cell
    return None

def getListOfCellsByCriteria(file: File, wsNumber=0, criteria, range: str):
    """Finds list of cells with values equal to criteria by
    searching in some range or in whole sheet

    Keyword arguments:
    criteria -- search criteria in cells values
    range -- search range (I22:J22, I, ...), by default set to None
                so it search in whole sheet
    """
    listOfCells = []

    if hasNumbers(range) == True:
        for cells in file.getWs(wsNumber)[range]:
            for cell in cells:
                if cell.value == criteria:
                    listOfCells.append(cell)
    elif hasNumbers(range) == False:
        for cell in file.getWs(wsNumber)[range]:
            if cell.value == criteria:
                listOfCells.append(cell)
    return listOfCells

def unmerge(file: File, wsNumber=0):
    for range in file.getWs(wsNumber).merged_cells.ranges:
        rangeList = list(range.bounds)
        minCol = rangeList[0]
        minRow = rangeList[1]
        maxCol = rangeList[2]
        maxRow = rangeList[3]
        file.getWs(wsNumber).unmerge_cells(start_row=minRow,
                                start_column=minCol,
                                end_row=maxRow,
                                end_column=maxCol
                                )
    return

def merge(file: File, wsNumber=0, range: str):
    start = range.split(":")[0]
    end = range.split(":")[1]
    minRow = openpyxl.utils.coordinate_to_tuple(start)[0]
    minCol = openpyxl.utils.coordinate_to_tuple(start)[1]
    maxRow = openpyxl.utils.coordinate_to_tuple(end)[0]
    maxCol = openpyxl.utils.coordinate_to_tuple(end)[1]
    file.getWs(wsNumber).merge_cells(start_row=minRow,
                        start_column=minCol,
                        end_row=maxRow,
                        end_column=maxCol
                        )
    return

def mergeByTuple(file: File, wsNumber=0, rangeList: list):
    for range in rangeList:
        coord = list(range.bounds)
        rangeStr = str(openpyxl.utils.get_column_letter(coord[0])) + str(
            coord[1]) + ":" + str(openpyxl.utils.get_column_letter(coord[2])) + str(coord[3])
        merge(rangeStr)
    return

def initHeader(file: File, wsNumber=0, headerRange: str):
    self.header = headerRange
    split = headerRange.split(":")

    try:
        if len(split) < 2:
            raise Exception()
    except Exception:
        print("Header can`t be 1 cell " + headerRange)

    self.headerDiapasone = self.ws[self.header]
    self.leftTopCoordinate = headerRange.split(":")[0]
    self.rightBotCoordinate = headerRange.split(":")[1]
    self.leftTopRow = openpyxl.utils.coordinate_to_tuple(self.leftTopCoordinate)[
        0]
    self.leftTopColumn = openpyxl.utils.coordinate_to_tuple(
        self.leftTopCoordinate)[1]
    self.rightBotRow = openpyxl.utils.coordinate_to_tuple(
        self.rightBotCoordinate)[0]
    self.rightBotColumn = openpyxl.utils.coordinate_to_tuple(
        self.rightBotCoordinate)[1]

def getHeadersOfAllColumns(self):
    """
    """
    listOfHeaders = []
    counter = 0
    for cells in self.headerDiapasone:
        for cell in cells:
            if cell.value != None:
                subList = [cell.value, cell.column, cell.row]
                listOfHeaders.append(subList)
                print(listOfHeaders[counter])
                counter += 1
    return listOfHeaders


if __name__ == "__main__":
    print("I`m excel.py file")
