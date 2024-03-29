import os
from shutil import copyfile

import openpyxl
import win32com.client
import pythoncom


def hasNumbers(inputString: str):
    """Checks string for containing numbers
    returns True if string has at least one digit
    """
    return any(char.isdigit() for char in inputString)


class File:

    def __init__(self, pathToFile: str, fileName: str):
        self.pathToFile = pathToFile
        self.fileName = fileName
        self.fileExtension = os.path.splitext(self.fileName)[1]
        self.fileNameWithoutExtension = os.path.splitext(self.fileName)[0]
        self.isOpened = False
        self.wasCalled = False
        self.shouldBeDeleted = True

    def open(self):
        pass

    def close(self):
        pass

    def save(self, path: str, name: str, extension=".xlsx"):
        pass

    def getWb(self):
        pass
    
    def getWs(self):
        pass

class PyWin(File):

    def open(self):
        """Opens file using pyWin. If it fails, then function will
        kill excel application
        """
        if self.isOpened == False:
            self.excelApp = win32com.client.Dispatch("Excel.Application")
            self.excelApp.Visible = False
            try:
                self.wb = self.excelApp.Workbooks.Open(
                    self.pathToFile + "\\" + self.fileName)
                self.isOpened = True
            except:
                self.isOpened = False
                self.excelApp.Quit()
                print("Программа не может открыть файл " + self.fileName)
                raise FileNotFoundError

    def getWb(self):
        """Returns workbook if file is opened
        """
        if self.isOpened == True:
            return self.wb

    def getWs(self, wsName="", isActiveSheet=False):
        """Returns worksheet if file is opened

        Keyword arguments:
        wsName -- name of worksheet.
        isActibeSheet -- If True then will return first 
                sheet or active sheet (sheet that was 
                opened last time)
        """
        if self.isOpened == True:
            if isActiveSheet == False:
                return self.wb.Worksheets(wsName)
            elif isActiveSheet == True:
                return self.wb.ActiveSheet 

    def getApp(self):
        """Returns instance of excel application 
        (if file was opened) for being able to 
        close app if needed
        """
        if self.isOpened == True:
            return self.excelApp

    def close(self):
        """Closes file and killing an excel app
        """
        if self.isOpened == True:
            self.wb.Close()
            
            self.excelApp.Quit()
            self.isOpened = False

    def save(self, path: str, name: str, extension=".xlsx", conflictResolution=False):
        """Saves file at path/name.xls or path/name.xlsx
        Keyword arguments:
        path -- full path to directory. Like C:\\User...(one slash
                instead of two should be used)
        name -- name of file without extension
        extension -- extension of excel file. Can be only .xls
                or .xlsx
        ConflictResolution -- (Not Working currently) if True then pyWin will overwrite file
                with similar name in directory 
        """
        if self.isOpened == True:
            if extension == ".xlsx":
                # Code for xslx format
                fileFormat = 51
            elif extension == ".xls":
                # Code for xls format
                fileFormat = 56
            if conflictResolution == True:
                self.excelApp.DisplayAlerts = False
                self.wb.SaveAs(path + "\\" + name + extension, 
                            FileFormat=fileFormat, ConflictResolution=2)
                self.excelApp.DisplayAlerts = True
            elif conflictResolution == False:
                self.wb.SaveAs(path + "\\" + name + extension, 
                            FileFormat=fileFormat)
            return True
        return False
        
    def insertColumn(self, column: str):
        """Incerts column using pyWin. 
        Function incerts column to the right at first worksheet

        Keyword arguments:
        column -- name of the column to the left of which the 
                new column will be inserted. Could be number 
                or letter like str(17) or "F"
        """
        ws = self.getWs(isActiveSheet=True)
        # select column as range object
        if column.isdigit():
            openpyxl.utils.get_column_letter(column)
            newColumn = openpyxl.utils.get_column_letter(column)
            rangeObj = ws.Range(newColumn+str(1)+str(":")+newColumn+str(2))
        else:
            rangeObj = ws.Range(column+str(1)+str(":")+column+str(2))
        rangeObj.EntireColumn.Insert()

    def insertRow(self, row: str):
        """Incerts row using pyWin. 
        Function incerts new row above the specified 
                at first worksheet

        Keyword arguments:
        row -- name of row above what the row would be inserted
        """
        ws = self.getWs(isActiveSheet=True)
        # select row as range object
        columnNumber = openpyxl.utils.get_column_letter(
                            ws.UsedRange.Columns.Count)   
        rangeObj = ws.Range("A"+row+":"+columnNumber+row)
        rangeObj.EntireRow.Insert()

class OpenPyXl(File):
    def open(self, data_only=True, keep_vba=False, keep_links=True, read_only=False):
        """Opens file using openpyxl. 

        Keyword arguments:
        data_only -- if True then only data without formulas
                would be in excel workbook
        """
        if self.isOpened == False:
            try:
                self.wb = openpyxl.load_workbook(self.pathToFile + "\\" + self.fileName,
                                                 data_only=data_only)
                self.isOpened = True
            except:
                self.isOpened = False
                print("Программа не может открыть файл " + self.fileName)
                raise FileNotFoundError

    def getWb(self):
        """Returns workbook if file is opened
        """
        if self.isOpened == True:
            return self.wb

    def getWs(self, wsName="", isActiveSheet=False):
        """Returns worksheet if file is opened

        Keyword arguments:
        wsName -- name of worksheet. If 0 or "" then 
                will return first sheet
        isActiveSheet -- If True then will return first 
                sheet
        """
        if self.isOpened == True:
            if isActiveSheet == True or wsName == 0 or wsName == "":
                return self.wb[self.getWsNames()[0]]
            elif isActiveSheet == False:
                return self.wb[wsName]

    def getWsNames(self):
        """Returns list of workbook sheetnames
        if file is opened
        """
        if self.isOpened == True:
            return self.wb.sheetnames

    def close(self):
        """Closes file if was opened
        """
        if self.isOpened == True:
            self.wb.close()
            self.isOpened = False

    def save(self, path: str, name: str, extension=".xlsx"):
        """Saves file at path/name.xls or path/name.xlsx
        If saving with .xls extension then will saves in .xlsx first
        and then will create temporary instance
        of pyWin file and  saves it with .xls extension

        Keyword arguments:
        path -- full path to directory. Like C:\\User...(one slash
                instead of two should be used)
        name -- name of file without extension
        extension -- extension of excel file. Can be only .xls
                or .xlsx
        """
        if self.isOpened == True:
            if extension == ".xlsx": 
                self.wb.save(path + "\\" + name + extension)
            elif extension == ".xls":
                self.wb.save(path + "\\" + name + ".xlsx")
                pyWinFile = PyWin(self.pathToFile, name + ".xlsx")
                pyWinFile.open()
                pyWinFile.save(self.pathToFile, name, extension=extension)
                pyWinFile.close()

    def setCellsInColumnByRowCoord(self, row: int, column: str, value, wsName=""):
        """Finds values in one column by row coordinate
        and then set value of that cell

        Keyword arguments:
        row -- row number
        column -- column in which the searh would happend
        value -- this would be set to the cell
        wsName -- name of worksheet. If empty then
                will work with first worksheet
        """
        for cells in self.getWs(wsName)[column]:
            for cell in cells:
                if cell.row == row:
                    cell.value = value
        return

    def setCellsInRowByColumnCoord(self, row: int, column: str, value, wsName=""):
        """Finds values in one row by column coordinate
        and then set value of that cell

        Keyword arguments:
        row -- row number
        column -- column in which the searh would happend
        value -- this would be set to the cell
        wsName -- name of worksheet. If empty then
                will work with first worksheet
        """
        for cells in self.getWs(wsName)[str(row)]:
            for cell in cells:
                if cell.column == openpyxl.utils.column_index_from_string(column):
                    cell.value = value
        return

    def getFirstCellByCriteria(self, criteria, range: str = None, wsName=""):
        """Finds first cell by searchin in whole sheet or in 
        some range the target criteria. Similar to Excel 
        function VLOOKUP (ВПР)

        Keyword arguments:
        criteria -- searching criteria (str, int, ...)
        range -- search range (I22:J22, or I), by default set to None
                    so it search in whole sheet
        wsName -- name of worksheet. If empty then
                will work with first worksheet
        """
        if range == None:
            for cells in self.getWs(wsName):
                for cell in cells:
                    if cell.value == criteria:
                        return cell
        elif range != None:
            if hasNumbers(range) == True:
                for cells in self.getWs(wsName)[range]:
                    for cell in cells:
                        if cell.value == criteria:
                            return cell
            elif hasNumbers(range) == False:
                for cell in self.getWs(wsName)[range]:
                    if cell.value == criteria:
                        return cell
        return None

    def getListOfCellsByCriteria(self, criteria, range: str, wsName=""):
        """Finds list of cells with values equal to criteria by
        searching in some range or in whole sheet

        Keyword arguments:
        criteria -- search criteria in cells values
        range -- search range (I22:J22, I, ...), by default set to None
                    so it search in whole sheet
        wsName -- name of worksheet. If empty then
                will work with first worksheet
        """
        listOfCells = []

        if hasNumbers(range) == True:
            for cells in self.getWs(wsName)[range]:
                for cell in cells:
                    if cell.value == criteria:
                        listOfCells.append(cell)
        elif hasNumbers(range) == False:
            for cell in self.getWs(wsName)[range]:
                if cell.value == criteria:
                    listOfCells.append(cell)
        return listOfCells

    def unmerge(self, wsName=""):
        """Unmerges all cells in worksheet

        Keyword arguments:
        wsName -- name of worksheet. If empty then
                will work with first worksheet
        """
        for range in self.getWs(wsName).merged_cells.ranges:
            rangeList = list(range.bounds)
            minCol = rangeList[0]
            minRow = rangeList[1]
            maxCol = rangeList[2]
            maxRow = rangeList[3]
            self.getWs(wsName).unmerge_cells(start_row=minRow,
                                    start_column=minCol,
                                    end_row=maxRow,
                                    end_column=maxCol
                                    )
        return

    def merge(self, range: str, wsName=""):
        """Merges range of cells

        Keyword arguments:
        range -- range of cells ("A1:G52")
        wsName -- name of worksheet. If empty then
                will work with first worksheet
        """
        start = range.split(":")[0]
        end = range.split(":")[1]
        minRow = openpyxl.utils.coordinate_to_tuple(start)[0]
        minCol = openpyxl.utils.coordinate_to_tuple(start)[1]
        maxRow = openpyxl.utils.coordinate_to_tuple(end)[0]
        maxCol = openpyxl.utils.coordinate_to_tuple(end)[1]
        self.getWs(wsName).merge_cells(start_row=minRow,
                            start_column=minCol,
                            end_row=maxRow,
                            end_column=maxCol
                            )
        return

class Manager:

    def __init__(self, pathToWorkDir: str):
        self.files = []
        self.pathToWorkDir = pathToWorkDir

    def addFileByPath(self, pathToFile: str, fileName: str, returnFile=False, defaultParser=True, openBy=0):
        """Adds file in manager class by path to this file

        Keyword argument:
        pathToFile -- path to directory with file, like C:\\User...(one slash
                        instead of two should be used)
        fileName -- name of file with extension
        returnFile -- if True then will return an appended file
                        if False then wouldn`t
        defaultParser -- if True then would decide by file extension
                        what class to use openpyxl or pyWin
                        if False then it should be specified in openBy 
                        variable
        openBy -- if defaultParser is True then would be passed
                        if defaultParser False then if:
                        openBy == 0 --> openpyxl
                        openBy == 1 --> pyWin
        """
        if defaultParser == True:
            if ".xlsx" in fileName:
                self.files.append(OpenPyXl(pathToFile, fileName))
            elif ".xls" in fileName:
                self.files.append(PyWin(pathToFile, fileName))
        elif defaultParser == False:
            if openBy == 0:
                self.files.append(OpenPyXl(pathToFile, fileName))
            elif openBy == 1:
                self.files.append(PyWin(pathToFile, fileName))
        if returnFile == True:
            self.files[len(self.files)-1].wasCalled = True
            return self.files[len(self.files)-1]

    def addFile(self, file, returnFile=False):
        """Adds file to manager

        Keyword parameters:
        file -- file instance
        returnFile -- if True then will return this
                    file fron manager
        """
        self.files.append(file)
        if returnFile == True:
            self.files[len(self.files)-1].wasCalled = True
            return self.files[len(self.files)-1]

    def addFilesInDir(self):
        """ Adds files in directory that
        was initialized when class instance was 
        created
        """
        for r, d, f in os.walk(self.pathToWorkDir):
            for fileName in f:
                self.addFileByPath(self.pathToWorkDir, fileName)
            break

    def __getitem__(self, i: int):
        #print("getitem:", i, len(self.files))
        if i < len(self.files) and i > -1:
            return self.files[i]
        else:
            raise AttributeError

    def removeFile(self, thatFile: File):
        """Removes file from manager
        Should be used like:
        manager.removeFile(manager.getFile(fileName))

        Keyword arguments:
        thatFile -- file instance that should
                    be removed
        """
        try:
            thatFile.close()
            self.files.remove(thatFile)
        except ValueError:
            print("Couldn`t remove file " + thatFile.fileName)
        # It`s not an error. It`s pylint problem
        except pythoncom.com_error:
            print("Couldn`t close file. COM exception occured")
            self.files.remove(thatFile)
        

    def deleteFile(self, thatFile: File, extension=".xlsx"):
        """Removes file from manager and 
        then deletes that file from system

        Keyword arguments:
        thatFile -- file instance, that would be deleted
        extension -- extension of file. By default set to
                        ".xlsx" but can be set to ".xls"
        """
        self.removeFile(thatFile)
        fullName = thatFile.pathToFile + "\\"
        fullName += thatFile.fileNameWithoutExtension
        fullName += thatFile.fileExtension
        os.remove(fullName)

    def getNumberOfFiles(self):
        """Returns number of file
        in manager
        """
        return len(self.files)

    def printAllFiles(self):
        """Prints all files in 
        manager with file name parameter and wasCalled 
        parameter and id
        """
        print("\n#----------#")
        for file in self.files:
            print("\tfile:")
            print("\t\t", "Name :", file.fileName)
            print("\t\t", "wasCalled :", file.wasCalled)
            print("\t\t", "isOpened :", file.isOpened)
            print("\t\t", "File :", file)
        print("#----------#\n")

    def getFile(self, partOfNameOfFile, extension=".xls", exactMatch=False):
        """Returns file by part of file 
        name and its extension

        Keyword arguments:
        partOfNameOfFile -- part of file name. If a file have
                a name "IAmTheFile.xls", then partOfNameOfFile
                could be "TheFile" or "iam", etc.
        extension -- extension of file. By default set to ".xls"
                but can be set to ".xlsx"
        exactMatch -- flag. If true then partOfNameOfFile should be 
                exactly as a file name
        """
        for file in self.files:
            if (exactMatch == False and partOfNameOfFile in file.fileName) or \
                            (exactMatch == True and \
                            partOfNameOfFile == file.fileNameWithoutExtension):
                if extension == file.fileExtension:
                    if file.wasCalled == False:
                        file.wasCalled = True
                        return file
        return None

    def removeUnCalledFiles(self):
        """Removes files that have wasCalled parameter
        set to False from manager
        """
        toRemove = []
        for file in self.files:
            if file.wasCalled == False:
                toRemove.append(file)
        # Remove duplicates by set()    
        for file in set(toRemove):
            self.removeFile(file)

    def deleteUnCalledFiles(self):
        """Removes from manager and then deletes
        from system files that have wasCalled
        parameter set to False
        """ 
        toDelete = []
        for file in self.files:
            if file.wasCalled == False and file.shouldBeDeleted:
                toDelete.append(file)
        # Remove duplicates by set()
        for file in set(toDelete):
            self.deleteFile(file)

    def deleteClosedFiles(self):
        """Removes from manager and then deletes
        from system files that have isOpened
        parameter set to False
        """
        forDelete = []
        for file in self.files:
            if file.isOpened == False and file.shouldBeDeleted:
                forDelete.append(file)

        # Remove duplicates by set()
        for file in set(forDelete):
            self.deleteFile(file)

    def allFromXlsToXlsx(self):
        """Resaves all files in manager in .xlsx
        format if they have .xls extension. After
        that it removes files with .xls extension
        from manager
        """
        forRemove = []
        for file in self.files:
            #print(file.fileName)
            if file.fileExtension == ".xls":
                file.open()
                file.save(file.pathToFile, file.fileNameWithoutExtension, ".xlsx")
                file.close()
                newFileName = file.fileNameWithoutExtension+".xlsx"
                self.addFileByPath(file.pathToFile, newFileName)
                file.wasCalled = False
                forRemove.append(file)
    
        for file in forRemove:
            self.removeFile(file)
        
        return

    def createDuplicate(self, file, duplicateName: str):
        """Creates duplicate of file in directory
        Duplicated file will have .xlsx extension

        Keyword arguments:
        file -- file that would be copied
        duplicateName -- name of duplicate file
        """
        copyfile(file.pathToFile + "\\" + file.fileName, 
                file.pathToFile + "\\" + duplicateName + ".xlsx")

        return


if __name__ == "__main__":
    print("I`m manager.py file")