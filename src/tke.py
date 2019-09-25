import os
import datetime
from copy import copy


import openpyxl
import win32com.client


from src.manager import *


class TKE:

    def __init__(self, dir: str):
        self.mng = Manager(os.path.abspath(dir))
        self.numberOfFilesToStart = 4
        self.checkIfDirectoryIsReady(dir)

    def checkIfDirectoryIsReady(self, path: str):

        self.mng.addFilesInDir()

        self.mng.getFile("Новый отчет")
        self.mng.getFile("Киiвтеплоенерго")
        self.mng.getFile("Звiт_Рестр")
        self.mng.getFile("90%ТКЕ_ПСО")

        self.mng.deleteUnCalledFiles()               
        self.mng.allFromXlsToXlsx()
        
        try:
            self.todayTKE = self.mng.getFile("Новый отчет", extension=".xlsx")
            self.yesterdayTKE = self.mng.getFile("90%ТКЕ_ПСО", extension=".xlsx")
            self.kyivEnergoPas = self.mng.getFile("Киiвтеплоенерго", extension=".xlsx")
            self.restructurization1730 = self.mng.getFile("Звiт_Рестр", extension=".xlsx")
            
            if self.mng.getNumberOfFiles() != self.numberOfFilesToStart:
                    raise AttributeError
        except AttributeError:
            print("Не хватает файлов для работы. Проверьте директорию " + str(path))
            msg = """Файлы, нужные для работы: 
            1. Новый отчет  свежая база ТКЕ_ПСО
            2. 90%ТКЕ_ПСО_... вчрашняя база ТКЕ_ПСО (там где ... там стоит дата с месяцем). 
                            Это файл, с которым будет сравниваться список предприятий (новички).
            3. Паспорт Киiвтеплоенерго КП ВО Киiвради (КМДА) за сегодня
            4. Звiт_Рестр_1730_Друк_ВсiОбластi  файл по 1730
            Итого: 4 экселевских файлов
            После исправления запустите программу заново. Сейчас программа завершит работу
            Нажмите любую клавишу а затем Enter
            """
            print(msg)
            input()
            exit()

    def deleteFiles(self, programmIsDone=True):
        """Deletes all created files with .xlsx extension
        """
        # If programm has daone its work then close files
        if programmIsDone == True:
            try:
                self.yesterdayTKE.close()
                self.kyivEnergoPas.close()
                self.restructurization1730.close()
            except:
                print("Программа не смогла закрыть экселевские файлы")
        self.mng.addFileByPath(self.todayTKE.pathToFile, 
                            self.generateName() + ".xlsx")
        self.mng.deleteClosedFiles()
        return

    def run(self):
        self.copyColumn()
        self.mainCalculations()
        name = self.generateName()
        self.hideColumns()
        self.addFilter()
        self.todayTKE.save(self.todayTKE.pathToFile, name, extension=".xls")
        self.deleteFiles()
        
    def mainCalculations(self):

        fileNameForRead = "forRead"
        self.mng.createDuplicate(self.todayTKE, fileNameForRead)
        #Read from/write to block
        # Write to
        self.todayTKE.open(data_only=False)
        todayWs = self.todayTKE.getWs("Sheet1")

        # Read from
        fileNameForRead+=".xlsx"
        todayTkeWithData = self.mng.addFileByPath(self.todayTKE.pathToFile, 
                            fileNameForRead, returnFile=True)
        todayTkeWithData.open(data_only=True)
        todayWsData = todayTkeWithData.getWs("Sheet1")

        numberOfRows = todayWs.max_row
        # Set 1 and 'договир э' to those companies who have a restructurization contract
        self.restructurization1730.open()
        rangeIter = "O12" + ":" + "O" + str(numberOfRows)
        for cells in todayWsData[rangeIter]:
            for cell in cells:
                if cell.value != "" and cell.value != None:
                    # Call restructurization() for get debt
                    summary = self.restructurization(todayWsData, cell.column, cell.row)
                    column = openpyxl.utils.column_index_from_string("AM")
                    if summary != None:
                        # If company have debt >0 then this summ will appear in
                        # column "AM"
                        todayWs.cell(column=column, row=cell.row).value = summary

                        columnWithPlan = openpyxl.utils.column_index_from_string("AU")
                        cellValueWithPlan = todayWs.cell(column=columnWithPlan, row=cell.row).value
                        # If there is no plan for this company, and this company have debp
                        # then set 0 to the column with conditions
                        if cellValueWithPlan == 0 or cellValueWithPlan == None or cellValueWithPlan == "":
                            columnWithConditions = openpyxl.utils.column_index_from_string("AT")
                            todayWs.cell(column=columnWithConditions, row=cell.row).value = 0
                    elif summary == None:
                        # If company have debt or debt <0 then "договір є" will
                        # appear in column "AM"
                        todayWs.cell(column=column, row=cell.row).value = str("договір є")
                    # This check needs for empty cell not to be filled
                    column = openpyxl.utils.column_index_from_string("AT")
                    if todayWsData.cell(column=column, row=cell.row).value != "":
                        todayWs.cell(column=column, row=cell.row).value = 1
        
        # Copy data from 'поточний лимит' to 'попередний лимит'
        rangeIter1 = "BO10" + ":" + "BU" + str(numberOfRows)
        rangeIter2 = "BB10" + ":" + "BH" + str(numberOfRows)
        for cells1, cells2 in zip(todayWsData[rangeIter1], todayWsData[rangeIter2]):
            for cell1, cell2 in zip(cells1, cells2):
                if cell1.row == cell2.row:
                    todayWs.cell(column=cell2.column, row=cell2.row).value = cell1.value

        # Set 'план э' to those rows wich have 0`s in both columns with conditions
        # Check the range
        list1 = todayTkeWithData.getListOfCellsByCriteria(0, "AS")  
        list2 = todayTkeWithData.getListOfCellsByCriteria(0, "AT")

        for cell1 in list1:
            if cell1.value == 0:
                for cell2 in list2:
                    if cell2.value == 0 and cell1.row == cell2.row:
                        column = openpyxl.utils.column_index_from_string("AU")
                        cellValueCheck = todayWsData.cell(column=column, row=cell1.row).value
                        # Check fo cells not to be empty
                        if cellValueCheck == 0 or cellValueCheck == None or cellValueCheck == "":
                            continue
                        else:
                            column = openpyxl.utils.column_index_from_string("AU")
                            todayWs.cell(column=column, row=cell1.row).value = str("план є")
                            todayWs.cell(column=column+1, row=cell1.row).value = ""
                            todayWs.cell(column=column+2, row=cell1.row).value = ""
                            todayWs.cell(column=column+3, row=cell1.row).value = ""
                            todayWs.cell(column=column+4, row=cell1.row).value = ""
                            todayWs.cell(column=column+5, row=cell1.row).value = ""
                            todayWs.cell(column=column+6, row=cell1.row).value = ""

        # Find the difference between columns with 'план на декаду' and 'поточний лимит'
        for row in range(10, numberOfRows):
            column = openpyxl.utils.column_index_from_string("AU")
            cellValueCheck = todayWsData.cell(column=column, row=row).value
            if cellValueCheck == None:
                continue
            else:
                columnAS = openpyxl.utils.column_index_from_string("AS")
                columnAT = openpyxl.utils.column_index_from_string("AT")
                cellValueCheck1 = todayWsData.cell(column=columnAS, row=row).value
                cellValueCheck2 = todayWsData.cell(column=columnAT, row=row).value
                # Check fo cells not to be empty
                if cellValueCheck1 == 0 and cellValueCheck2 == 0:
                    continue
                elif cellValueCheck1 == None and cellValueCheck2 == None:
                    continue
                else:
                    column1 = openpyxl.utils.column_index_from_string("BO")
                    value1 = todayWsData.cell(column=column1, row=row).value
                    value2 = todayWsData.cell(column=column, row=row).value
                    dx = value1 - value2
                    if dx > 1e-6 or dx < -1e-6:
                        column = openpyxl.utils.column_index_from_string("BW")
                        todayWs.cell(column=column, row=row).value = dx

        self.kyivEnergoMoney(todayTkeWithData)
        self.garantMM(todayTkeWithData)
        todayTkeWithData.close()
        return

    def copyColumn(self):

        # Save to tmp variable today and yesterday TKE files
        tmpTodayTKE = self.todayTKE
        tmpYesterdayTKE = self.yesterdayTKE
        # Opens today and yesterday TKE files with pyWin
        self.todayTKE = self.mng.addFileByPath(self.todayTKE.pathToFile, 
                        self.todayTKE.fileName, returnFile=True, defaultParser=False, openBy=1)
        self.yesterdayTKE = self.mng.addFileByPath(self.yesterdayTKE.pathToFile, 
                        self.yesterdayTKE.fileName, returnFile=True, defaultParser=False, openBy=1)
        self.todayTKE.open()
        self.yesterdayTKE.open()
        # Set first sheet as active
        todayWs = self.todayTKE.getWs("Sheet1")
        yestWs = self.yesterdayTKE.getWs("Sheet1")
        # Looks through all rows in today TKE and compare values in "R"
        # column (wich corresponds to company name) and if values don`t match
        # then it say`s that there is a new company in today TKE and it should 
        # be copied to yesterday TKE
        column = openpyxl.utils.column_index_from_string("R")
        numberOfCycles = 0
        while True:
            wasMismatch = False
            for row in range(10, todayWs.UsedRange.Rows.Count):
                value1 = todayWs.Cells(row, column).Value
                value2 = yestWs.Cells(row, column).Value
                if value1 != value2:
                    wasMismatch = True
                    print("Внимание!!! Новое предприятие:", value1)
                    self.yesterdayTKE.insertRow(str(row))
                    for column1 in range(1, yestWs.UsedRange.Columns.Count):
                        yestWs.Cells(row, column1).Value = todayWs.Cells(row, column1).Value
            if wasMismatch == False:
                break
            elif wasMismatch == True:
                numberOfCycles += 1
            if numberOfCycles > 5:
                print("Внимание!!! Слишком много несовпадений предприятий со вчерашним днем")
                print("Возможна ошибка")
                
        # Incerts column left to "AS" column in today TKE and then copies column 
        # "AS" from yesterday TKE and incerts it to created column in today TKE
        self.todayTKE.insertColumn("AS")
        todayWs.Range("AS1:AS2").EntireColumn.Unmerge()
        yestWs.Range("AS1:AS2").EntireColumn.Unmerge()
        yestWs.Range("AS1:AS2").EntireColumn.Copy()
        todayWs.Paste(todayWs.Range("AS1:AS2"))
        # Saves files with rewriting exsited files in directory
        self.todayTKE.save(self.todayTKE.pathToFile, self.todayTKE.fileNameWithoutExtension)
        self.yesterdayTKE.save(self.yesterdayTKE.pathToFile, self.yesterdayTKE.fileNameWithoutExtension)
        self.todayTKE.close()
        self.mng.removeUnCalledFiles()
        # Returns tmp files to variables
        self.todayTKE = tmpTodayTKE
        self.yesterdayTKE = tmpYesterdayTKE

        return

    def restructurization(self, ws, column: int, row: int):
        """Looks through 1730 file and finds company`s debt
        Returns summary debt if its >0, and None if <0

        Keyword arguments:
        ws -- today TKE worksheet
        column -- column company number
        row -- row company number
        """
        # Set EGRPOU value wich is placed in column right next to 
        # column value
        EGRPOU = ws.cell(column=column+1, row=row).value
        try:
            wsRestr = self.restructurization1730.getWs("Sheet1")
            # Get cell with EGRPOU value in 1730 file
            row = self.restructurization1730.getFirstCellByCriteria(EGRPOU, "D").row-1
        except AttributeError:
            print("В списках договоров реструктуризации " + \
                        "1730 не найдено предприятие с кодом ЕГРПОУ", EGRPOU)
            return None
        overpaymentColumn = openpyxl.utils.column_index_from_string("U")
        debtColumn = openpyxl.utils.column_index_from_string("V")
        overpayment = wsRestr.cell(column=overpaymentColumn, row=row).value
        debt = wsRestr.cell(column=debtColumn, row=row).value
        # Summary of debt (wich are positive values) and overpayment (negative value)
        summary = overpayment + debt
        if summary > 0:
            return summary
        if summary <= 0:
            return None

    def kyivEnergoMoney(self, dataFile):
        """Finds kyiv teplo energo money in their passport file
        Keyword arguments:
        dataFile -- TKE file
        """
        try:
            self.kyivEnergoPas.open()
            kyivWs = self.kyivEnergoPas.getWs("Sheet1")
            
            headerColumn = openpyxl.utils.column_index_from_string(str("A"))
            contractColumn = openpyxl.utils.column_index_from_string(str("B"))
            moneyColumn = openpyxl.utils.column_index_from_string(str("I"))
            # Get list with cells with values "Період" in first column ("A")
            listOfHeaders = self.kyivEnergoPas.getListOfCellsByCriteria("Період", "A")
            # Take row of second cell
            row = listOfHeaders[1].row
            # From that row iterate untill cell value would contain "рік" and
            # if column next to column "A" not contain "РЗ" in its contract then
            # value in column "I" would be summed up
            money = 0
            while True:
                row += 1
                header = kyivWs.cell(column=headerColumn, row=row).value
                if header == "" or header == None:
                    break
                if "рік" in header:
                    contract = kyivWs.cell(column=contractColumn, row=row).value
                    if "РЗ" not in contract:
                        money += kyivWs.cell(column=moneyColumn, row=row).value
            
            self.kyivEnergoPas.close()
        except:
            money = 0
            print("Проблема с подсчетом оплаты Київтеплоенерго КП ВО")
        
        try:
            # Finds cell with "Київтеплоенерго КП ВО Київради (КМДА)" in TKE file
            kyivEnergoRow = dataFile.getFirstCellByCriteria("Київтеплоенерго " + \
                            "КП ВО Київради (КМДА)", "R").row
            ws = self.todayTKE.getWs("Sheet1")

            # Payment column value
            column=openpyxl.utils.column_index_from_string(str("AQ"))
            ws.cell(column=column, row=kyivEnergoRow).value = money
            # For all contracts column value
            column=openpyxl.utils.column_index_from_string(str("AE"))
            column1=openpyxl.utils.column_index_from_string(str("AF"))
            column2=openpyxl.utils.column_index_from_string(str("AD"))
            # Set the right value in cell with debt
            ws.cell(column=column, row=kyivEnergoRow).value = \
                            ws.cell(column=column1, row=kyivEnergoRow).value - \
                            ws.cell(column=column2, row=kyivEnergoRow).value - money
        except:
            print("Программа не смогла внести данные о задолженности Київтеплоенерго КП ВО")
        
        return 

    def garantMM(self, dataFile):
        """Set right calculation of debt for garant energo MM
        Keyword arguments:
        dataFile -- TKE file
        """
        try:
            garantRow = dataFile.getFirstCellByCriteria("Гарант Енерго М ПП", "R").row
            ws = self.todayTKE.getWs("Sheet1")
            # For all contracts column value
            column=openpyxl.utils.column_index_from_string(str("AE"))
            column1=openpyxl.utils.column_index_from_string(str("AF"))
            # Payment column value
            column2=openpyxl.utils.column_index_from_string(str("AQ"))
            ws.cell(column=column, row=garantRow).value = \
                            ws.cell(column=column1, row=garantRow).value - \
                            ws.cell(column=column2, row=garantRow).value
        except:
            print("Программа не смогла внести данные о задолженности Гарант Енерго М ПП")

    def hideColumns(self):
        self.yesterdayTKE.open()
        # Get list of hidden colulmns
        listOfHiddenColumns = []
        for column in range(1, self.yesterdayTKE.getWs().max_column):
            columnLetter = openpyxl.utils.get_column_letter(column)
            isHidden = self.yesterdayTKE.getWs().column_dimensions[columnLetter].hidden
            listOfHiddenColumns.append(isHidden)
        # Hide columns in today TKE by list
        for column in range(1, self.todayTKE.getWs().max_column):
            if column < len(listOfHiddenColumns):
                if listOfHiddenColumns[column] == True:
                    columnLetter = openpyxl.utils.get_column_letter(column)
                    self.todayTKE.getWs().column_dimensions[columnLetter].hidden = True
        # Additional hiding
        listOfHiddenColumns = ["M", "O", "AW", "AX", "AY", 
                                "AZ", "BA", "BB", "BC", "BD", 
                                "BE", "BF", "BG", "BH", "BI",
                                "BJ", "BK", "BL", "BM", "BN",
                                "BP", "BQ", "BR", "BS", "BT",
                                "BU"]
        for columnLetter in listOfHiddenColumns:
            self.todayTKE.getWs().column_dimensions[columnLetter].hidden = True
        
        return

    def addFilter(self):
        FullRange = "A10:" + openpyxl.utils.get_column_letter(
                                self.todayTKE.getWs().max_column) + \
                                str(self.todayTKE.getWs().max_row)
        self.todayTKE.getWs().auto_filter.ref = FullRange

    def generateName(self):
        """Generates name for file TKE_ПСО
        with current date and month
        Returns string without file extension
        """
        day = datetime.datetime.today().day
        month = datetime.datetime.today().month
        year = datetime.datetime.today().year
        monthInRussian = [r"январь", r"февраль", r"март", 
                            r"апрель", r"май", r"июнь", 
                            r"июль", r"август", r"сентябрь", 
                            r"октябрь", r"ноябрь", r"декабрь"]
        fileName = "90%ТКЕ_ПСО_" + monthInRussian[month-1]
        fileName += "(" + str(day) + "."
        if month < 10:
            fileName += "0" + str(month) + "."
        elif month >= 10:
            fileName += str(month) + "."
        fileName += str(year) + ")"
        return fileName


