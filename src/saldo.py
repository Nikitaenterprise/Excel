import openpyxl
import win32com.client

def findInSaldo(saldoSheet, whatToFind: str, 
                whatCategory: list, whatResource: list, 
                whatColumns: list, inWhatColumnFind="A"):
    """Search saldo excel sheet for company name or other 
    value equal to whatToFind in column inWhatColumnFind.
    Returns list with size of whatColumns list size.

    Keyword arguments:
    saldoSheet -- excel file with saldo
    whatToFind -- criteria to find in column inWhatColumnFind
    whatCategory -- list with categories. 
                    ["TE", "КП"] will return all values 
                                    with this criteria
                    ["!TE"] will return all values except "TE"
    whatResource -- list with resource. 
                    ["2018", "2019"] will return all values 
                                    with this criteria
                    ["!2019"] will return all values except "2019"
    whatColumns -- list of columns from which the data would be taken
    inWhatColumnFind -- column in which the searching of criteria would
                        be happening
    """
    rangeIter = inWhatColumnFind + "10" + ":" +\
                inWhatColumnFind + str(saldoSheet.max_row)
    columnCategory = openpyxl.utils.column_index_from_string("C")
    columnResource = openpyxl.utils.column_index_from_string("F")

    listOfColumns = []
    for column in whatColumns:
        col = openpyxl.utils.column_index_from_string(column)
        listOfColumns.append(col)

    additionListCategory = []
    exclusionListCategory = []
    if whatCategory != None:
        for cat in whatCategory:
            if "!" in cat:
                exclusionListCategory.append(cat.split("!")[1])
            elif "!" not in cat:
                additionListCategory.append(cat)
    

    additionListResource = []
    exclusionListResource = []
    if whatResource != None:
        for res in whatResource:
            if "!" in res:
                exclusionListResource.append(res.split("!")[1]) 
            elif "!" not in res:
                additionListResource.append(res)
    

    for cells in saldoSheet[rangeIter]:
        for cell in cells:
            # If company name equals to what to find variable
            if cell.value != None and cell.value == whatToFind:
                # Iterate through this company data
                row = cell.row

                returnValuesList = [0]*len(whatColumns)
                while True:
                    # Move 1 row down
                    row += 1
                    category = saldoSheet.cell(column=columnCategory,
                                                row=row).value
                    # If its None then we know that company data ends
                    # and other company begins
                    if category == None:
                        break
                    # Transform int value of resource into str 2019 -> "2019"
                    resource = str(saldoSheet.cell(column=columnResource,
                                                    row=row).value).strip()
                    
                    # Get values from different columns
                    # Write them to list
                    valuesList = []
                    for column in listOfColumns:
                        value = saldoSheet.cell(column=column,
                                                row=row).value
                        if value == None:
                            value = 0
                        valuesList.append(value)
                    
                    willBeCalculatedCategory = False
                    willBeCalculatedResource = False

                    willBeCalculatedCategory = makeDecision(
                                additionListCategory,
                                exclusionListCategory,
                                category)
                    willBeCalculatedResource = makeDecision(
                                additionListResource,
                                exclusionListResource,
                                resource)
                    
                    if (willBeCalculatedCategory and 
                        willBeCalculatedResource):
                        for i in range(0, len(valuesList)):
                            returnValuesList[i] += valuesList[i]
                
                return returnValuesList

    # If no data was found then return zeros
    returnValuesList = [0]*len(listOfColumns)
    return returnValuesList

def findInSaldoAllValues(saldoSheet, whatCategory: list, 
                        whatResource: list, whatColumns: list):
    """Search saldo excel sheet for all values with some criteria 
    Returns list with size of whatColumns list size.

    Keyword arguments:
    saldoSheet -- excel file with saldo
    whatCategory -- list with categories. 
                    ["TE", "КП"] will return all values 
                                    with this criteria
                    ["!TE"] will return all values except "TE"
    whatResource -- list with resource. 
                    ["2018", "2019"] will return all values 
                                    with this criteria
                    ["!2019"] will return all values except "2019"
    whatColumns -- list of columns from which the data would be taken
    """
    rangeIter = "A10" + ":" + "A" + str(saldoSheet.max_row)
    columnCategory = openpyxl.utils.column_index_from_string("C")
    columnResource = openpyxl.utils.column_index_from_string("F")
    
    listOfColumns = []
    for column in whatColumns:
        col = openpyxl.utils.column_index_from_string(column)
        listOfColumns.append(col)

    additionListCategory = []
    exclusionListCategory = []
    if whatCategory != None:
        for cat in whatCategory:
            if "!" in cat:
                exclusionListCategory.append(cat.split("!")[1])
            elif "!" not in cat:
                additionListCategory.append(cat)
    

    additionListResource = []
    exclusionListResource = []
    if whatResource != None:
        for res in whatResource:
            if "!" in res:
                exclusionListResource.append(res.split("!")[1]) 
            elif "!" not in res:
                additionListResource.append(res)

    returnValuesList = [0]*len(whatColumns)

    for cells in saldoSheet[rangeIter]:
        for cell in cells:
            
            category = saldoSheet.cell(column=columnCategory,
                                    row=cell.row).value
            # Transform int value of resource into str 2019 -> "2019"
            resource = str(saldoSheet.cell(column=columnResource,
                                    row=cell.row).value).strip()
            
            # Next row if this is name of company
            if category == None and resource == "None":
                continue

            # Get values from different columns
            # Write them to list
            valuesList = []
            for column in listOfColumns:
                value = saldoSheet.cell(column=column,
                                        row=cell.row).value
                if value == None:
                    value = 0
                valuesList.append(value)
                
            willBeCalculatedCategory = False
            willBeCalculatedResource = False

            willBeCalculatedCategory = makeDecision(
                                additionListCategory,
                                exclusionListCategory,
                                category)
            willBeCalculatedResource = makeDecision(
                                additionListResource,
                                exclusionListResource,
                                resource)

            if (willBeCalculatedCategory and 
                willBeCalculatedResource):
                for i in range(0, len(valuesList)):
                    returnValuesList[i] += valuesList[i]

    return returnValuesList


def makeDecision(addList: list, excudeList: list, value):
    
    boolValue = False

    if (not addList and
        not excudeList):

        boolValue = True

    elif (addList and
        not excudeList and
        value in addList):

        boolValue = True

    elif (not addList and
        excudeList and
        value not in excudeList):
        
        boolValue = True

    elif (addList and
        excudeList):

        boolValue = True

    return boolValue