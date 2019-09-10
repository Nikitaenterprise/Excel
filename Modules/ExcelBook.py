import os

import openpyxl
import win32com


class ExcelBook:


    def __init__(self, name: str, data_only=True, read=True, worksheet: int=0):
        self.nameOfFile = name
        self.data_only = data_only
        if read == True:
            self.readExcelFile()

    def readExcelFile(self):
        """Reads file after creation of class instance by default.
        If this function wasn`t called after initialization of class instance
        (for example if you dont want to allocate a lot of memory by opening
        many excel books, but you want to initialize a class instances)
        then it can be called after.
        """
        extension = os.path.splitext(self.nameOfFile)[1] # may be .xls or .xlsx
        if extension == ".xls":
            self.reSaveFromXlsToXlsx(self.nameOfFile)
            extension = os.path.splitext(self.nameOfFile)[1] # reinitialize extension
        if extension == ".xlsx":
            self.wb = openpyxl.load_workbook(self.nameOfFile, data_only=self.data_only)
            self.ws = self.wb[self.wb.sheetnames[0]]
            self.vv = 2
        return True

    def reSaveFromXlsToXlsx(self, name: str):
        """Opens file in .xls format and saves it 
        in .xlsx format using pyWin32
        
        Keyword arguments:
        name -- name of file that will be changed
        """
        newName = os.path.splitext(name)[0]
        newName += str(".xlsx")
        excelApp = win32com.client.Dispatch("Excel.Application")
        excelApp.Visible = False
        wb = excelApp.Workbooks.Open(os.path.abspath(name))
        xlsx = 51 # Code for xslx format
        wb.SaveAs(os.path.abspath(newName), FileFormat=xlsx)
        self.nameOfFile = newName
        excelApp.Quit()
        return

    def save(self, name: str):
        """Closes file and save it to project root dir
        with a given name

        Keyword arguments:
        name -- name of saved file
        """
        self.wb.save(name)
        self.wb.close()
        return
    
    def close(self):
        """Closes file without saving"""
        self.wb.close()
        return 

    def generator(self, range: str):
        """Generator through range in sheet.
        Returns cell
        """
        for cells in self.ws[range]:
            for cell in cells:
                yield cell

    def setCellsInColumnByRowCoord(self, row: int, column: str, value):
        """Finds values in one column by row coordinate
        and then set value of that cell

        Keyword arguments:
        row -- row number
        column -- column in which the searh would happend
        value -- this would be set to the cell
        """
        for cells in self.ws[column]:
            for cell in cells:
                if cell.row == row:
                    cell.value = value
        return

    def setCellsInRowByColumnCoord(self, row: int, column: str, value):
        """Finds values in one row by column coordinate
        and then set value of that cell

        Keyword arguments:
        row -- row number
        column -- column in which the searh would happend
        value -- this would be set to the cell
        """
        for cells in self.ws[str(row)]:
            for cell in cells:
                if cell.column == openpyxl.utils.column_index_from_string(column):
                    cell.value = value
        return

    def findCellByStr(self, value, range: str = None):
        """Finds first cell by searchin in whole sheet 
        the target value

        Keyword argument:
        value -- searching value (str, int, ...)
        range -- search range (I22:J22, or I), by default set to None
        """
        if range == None:
            diapason = self.ws
        if range != None:
            diapason = self.ws[range]

        for cell in diapason:
            if cell.value == value:
                return cell
        return None

    def getListOfCellsWithCriteria(self, range: str, criteria):
        """Returns list of cells with values equal to criteria

        Keyword argument:
        range -- search range (I22:J22, I, ...)
        criteria -- search criteria in cells values
        """
        listOfCells = []

        if  range.isdigit() == True:
            for cells in self.ws[range]:
                for cell in cells:
                    if cell.value == criteria:
                        listOfCells.append(cell)
        elif range.isdigit() == False:
            for cell in self.ws[range]:
                if cell.value == criteria:
                    listOfCells.append(cell)
        return listOfCells

    def checkRowForRegion(self, row: int, column: int):
        """In excel files sometimes it can be observed that some rows
        contains sum of data of other rows below or under them.
        Such rows often contain name of 'область' on their left column
        Under them there are companies in this district.
        This function checks the row for containing name of district
        in specific column

        TODO

        Keyword arguments:
        range
        """
        return

    def unmerge(self):
        for range in self.ws.merged_cells.ranges:
            rangeList = list(range.bounds)
            minCol = rangeList[0]
            minRow = rangeList[1]
            maxCol = rangeList[2]
            maxRow = rangeList[3]
            self.ws.unmerge_cells(start_row=minRow,
                                start_column=minCol,
                                end_row=maxRow,
                                end_column=maxCol
                                )
        return


    def merge(self, range: str):
        start = range.split(":")[0]
        end = range.split(":")[1]
        minRow = openpyxl.utils.coordinate_to_tuple(start)[0]
        minCol = openpyxl.utils.coordinate_to_tuple(start)[1]
        maxRow = openpyxl.utils.coordinate_to_tuple(end)[0]
        maxCol = openpyxl.utils.coordinate_to_tuple(end)[1]
        self.ws.merge_cells(start_row=minRow,
                            start_column=minCol,
                            end_row=maxRow,
                            end_column=maxCol
                            )
        return

    def mergeByTuple(self, rangeList: list):
        for range in rangeList:
            coord = list(range.bounds)
            rangeStr = str(openpyxl.utils.get_column_letter(coord[0])) + str(coord[1]) + ":" + str(openpyxl.utils.get_column_letter(coord[2])) + str(coord[3])
            self.merge(rangeStr)
        return


if __name__ == "__main__":
    print("I`m ExcelBook.py file")
